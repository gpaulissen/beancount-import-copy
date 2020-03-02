# -*-coding: utf-8-*-
# coding=utf-8

"""ICScards.nl transaction and balance source.

Data format
===========

To use, first download transaction (and balance) data into a directory on the
filesystem.  The easiest way to download data from ICScards in the requisite format
is to use the finance_dl.icscards module. 

The transactions files are in PDF format and although it is possible to use
Python to create a CSV file from it (for example using the pdftotext script as
in other beanimport modules), the result is not at all good enough to use
it. I therefore suggest that you install 'WPS PDF to Word' (also converts to
Excel XLSX) from https://wps.com and later use Python module openpyxl to read it.

So the suggested workflow is:
1) Download the transaction statements from https://icscards.nl either
manually or automatically.
2) Use 'WPS PDF to Word' to create XLSX files.
3) Place those XLSX files in the data directory structure (see below).
4) Then this beancount_import module will take care of the rest.

You might have a directory structure like:

    financial/
      data/
        icscards/
          account_id/
            Rekeningoverzicht-54280230027-2020-01.xlsx


The `Rekeningoverzicht-54280230027-2020-01.xlsx` file should be a XLSX file
containing all downloaded transactions, in the normal PDF download format
provided by ICScards.  See the `testdata/source/icscards` directory for an
example.

The XLSX (converted to CSV for your convenience because including an XLSX in
source is not very helpful) should be of the form:

Line Text

1    "International Card Services BV Postbus 23225
     1100 DS Diemen
     Telefoon 020 - 6 600 600
     Kvk Amsterdam nr. 33.200.596",,,,,"www.icscards.nl
     Bankrek. NL99ABNA9999999999 BIC: ABNANL2A
     ICS identificatienummer bij incasso: NL99ZZZ999999999999",,,,,,,
2    Datum                                          ICS-klantnummer                         Volgnummer                                Bladnummer,,,,,,,,,,,,
3    17 januari 2020                           99999999999                               1                                                   1  van 1,,,,,,,,,,,,
4    "Vorig openstaand saldo               Totaal ontvangen betalingen       Totaal nieuwe uitgaven                Nieuw openstaand saldo
     € 1.801,55                          Af      € 1.827,97                          Bij      € 1.930,18                           Af     € 1.903,76                          Af",,,,,,,,,,,,
5    "Datum
     transactie",,"Datum
     boeking",,"Omschrijving                                                                                           Bedrag in
     vreemde valuta",,,,,,"Bedrag
     in euro's",,
6    24 dec,24 dec,,"IDEAL BETALING, DANK U",,,,,1000,,,Bij,
7    05 jan,05 jan,,"IDEAL BETALING, DANK U",,,,,801.55,,,Bij,
8    "Uw Card met als laatste vier cijfers 0467
     G.J.L.M. PAULISSEN",,,,,,,,,,,,
9    16 dec,18 dec,,SARL THONIC,,,ZZZZ,FR,,,65.00,Af,
10   16 dec,18 dec,,SODEXO,,,ZZZZZZZZZZZZZ,FR,,,10.00,Af,
11   18 dec,19 dec,,OW125692REDHIP12,,,ZZZZZZZZZZZZZ,FR,,,32.00,Af,
12   18 dec,19 dec,,SODEXO,,,ZZZZZZZZZZZZZ,FR,,,10.00,Af,
13   18 dec,19 dec,,HOTEL MERCURE,,,ZZZZZZZZZZZZZ,FR,,,422.05,Af,
14   18 dec,20 dec,,G7,,,ZZZZZZ,FR,,,50.00,Af,
15   20 dec,21 dec,,NETFLIX.COM,,,999-999-9999,NL,,,7.99,Af,
16   23 dec,24 dec,,SPA,,,ZZZZZZ,BE,,,76.20,Af,
17   25 dec,27 dec,,RESTAURANT,,,ZZZZZ,NL,,,475.00,Af,
18   26 dec,27 dec,,FLETCHER HOTEL,,,ZZZZZZZZZZ,NL,,,154.45,Af,
19   02 jan,03 jan,,QUICKEN INC,,,9999999999,US,,29.990000000,27.41,Af,
20   ,,,Wisselkoers USD,,,1.09413,,,,,,
21   08 jan,09 jan,,SODEXO,,,ZZZZZZZZZZZZZ,FR,,,10.00,Af,
22   08 jan,09 jan,,HOTEL MERCURE,,,ZZZZZZZZZZZZZ,FR,,,359.34,Af,
23   08 jan,09 jan,,UBER TRIP HELP.UBER.COM,,,ZZZZZZZZZZZZZ,FR,,,49.59,Af,
24   08 jan,09 jan,,SNCF,,,ZZZZZZZZZZ,FR,,,6.15,Af,
25   10 jan,10 jan,,QUICKEN INC,,,9999999999,US,,29.990000000,26.42,Bij,
26   ,,,Wisselkoers USD,,,1.13512,,,,,,
27   13 jan,14 jan,,SNCF OUIGO,,,ZZZZZZZ,FR,,,45.00,Af,
28   14 jan,15 jan,,SNCF WEB MOBILE,,,ZZZZZZZZ,FR,,,130.00,Af,
29   "Uw betalingen aan International Card Services BV zijn bijgewerkt tot 17 januari 2020.
     Het minimaal te betalen bedrag ad € 1.903,76 verwachten wij voor 7 februari 2020 op rekening NL99 ABNA 9999 9999 99 t.n.v. ICS in Diemen. Vermeld bij uw betaling altijd uw ICS-klantnummer 99999999999.",,,,,,,,,,,,
30   Bestedingslimiet                                                Minimaal te betalen bedrag,,,,,,,,,,,,
31   "€ 2.500                                                               € 1.903,76",,,,,,,,,,,,
32   Dit product valt onder het depositogarantiestelsel. Meer informatie vindt u op www.icscards.nl/depositogarantiestelsel en op het informatieblad dat u jaarlijks ontvangt.,,,,,,,,,,,,

The first line is only shown on the first sheet and contains some general information about ICScards.

The associated account is specified in row 3, column 2 below heading
'ICS-klantnummer' (ICS client number).

Lines 2 and 3 are repeated every page (a new sheet in the XLSX) and show the
page number.

The balance is in the fourth line in the first column (actually a list
containing 12 columns separated by at least two spaces or a newline).

The payments to your ICScards card are in lines 6 and 7, 5 non-empty columns in
total.

The payments from your ICScards are in line 9 till 28 where lines 20 and 26 are
just information about currency conversions. Please note that line 25 contains
a total cost in USD although the CSV does not show it (but the XLSX does).  The
number of columns is 7 or 8 dependening on whether the cost is included.

Specifying the source to beancount_import
=========================================

Within your Python script for invoking beancount_import, you might use an
expression like the following to specify the icscards source:

    dict(module='beancount_import.source.icscards',
         filenames=(
             glob.glob(os.path.join(journal_dir, 'data/icscards/*/*.xlsx'))
         ),
    )

where `journal_dir` refers to the financial/ directory.

Imported transaction format
===========================

If you receive a payment, or make a payment from your ICScards balance, a single
transaction of the following form is generated:

    2017-09-06 * "Sally Smith" "Rent"
      Assets:ICScards     1150.00 USD
        source_desc: "Rent"
      Expenses:FIXME  -1150.00 USD

"""

from typing import List, Union, Optional, Set
from openpyxl import load_workbook
from openpyxl.cell.read_only import ReadOnlyCell
import datetime
import collections
import re
import os
import locale
import traceback
from math import copysign

from beancount.core.data import Transaction, Posting, Balance, EMPTY_SET
from beancount.core.amount import Amount
from beancount.core.flags import FLAG_OKAY
from beancount.core.number import MISSING, D, ZERO

from . import description_based_source
from . import ImportResult, SourceResults
from ..matching import FIXME_ACCOUNT
from ..journal_editor import JournalEditor

DEBUG = False

def convert_str_to_list(str, max_items, sep=r'\s\s+|\t|\n'):
    return [x for x in re.split(sep, str)[0:max_items]]

# account may be either the icscards_id or the journal account name
ICScardsEntry = collections.namedtuple(
    'ICScardsEntry',
    ['account', 'date', 'amount', 'cost', 'payee', 'narration', 'source_desc', 'filename', 'line'])
RawBalance = collections.namedtuple(
    'RawBalance', ['account', 'date', 'amount', 'filename', 'line'])

def get_info(raw_entry: Union[ICScardsEntry, RawBalance]) -> dict:
    return dict(
        type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=raw_entry.filename,
        line=raw_entry.line,
    )

def load_transactions(filename: str, currency: str = 'EUR') -> [List[ICScardsEntry], List[RawBalance]]:
    def add_years(d, years):
        """Return a date that's `years` years after the date (or datetime)
        object `d`. Return the same calendar date (month and day) in the
        destination year, if it exists, otherwise use the following day
        (thus changing February 29 to March 1).

        """
        try:
            return d.replace(year = d.year + years)
        except ValueError:
            return d + (date(d.year + years, 1, 1) - date(d.year, 1, 1))

    def get_date(s: str):
        d = datetime.datetime.strptime(s, '%d %b').date()
        # Without a year it will be 1900 so augment
        while d <= page_date:
            d = add_years(d, 1)
        return add_years(d, -1)

    def get_amount(sheet_in, line_in, amount_in, transaction_type_in, foreign_currency_in=False):
        sign_out, amount_out, currency_out = 1, None, 'EUR'

        # determine sign_out
        if isinstance(transaction_type_in, str):
            if transaction_type_in == 'Af':
                sign_out = -1
            elif transaction_type_in == 'Bij':
                pass
            else:
                raise RuntimeError('Unknown transaction type "{0}" in sheet {1}, row {2}'.format(transaction_type_in, sheet_in, line_in))
        elif isinstance(transaction_type_in, ReadOnlyCell):
            if transaction_type_in.value == 'Af':
                sign_out = -1
            elif transaction_type_in.value == 'Bij':
                pass
            else:
                raise RuntimeError('Unknown transaction type "{0}" in sheet {1}, row {2}'.format(transaction_type_in.value, sheet_in, line_in))
        else:
            raise RuntimeError('Wrong type for transaction type "{0}" in sheet {1}, row {2}'.format(type(transaction_type_in), sheet_in, line_in))
                
        # determine amount_out and currency_out
        if isinstance(amount_in, str):
            # Is amount something like 1.827,97 ? (dutch) or 1,827.97 ?
            if amount_in[-3] == ',':
                amount_in = amount_in.replace('.', '').replace(',', '.')
            # amount something like € 1.827,97 ?
            if not(amount_in[0] >= '0' and amount_in[0] <= '9') and amount_in[1] == ' ':
                # Skip euro sign and take care of comma's and points
                amount_out = amount_in[2:]
            else:
                # Take care of comma's and points
                amount_out = amount_in                
        elif isinstance(amount_in, ReadOnlyCell) and foreign_currency_in:
            if isinstance(amount_in.value, str):
                # '235,01 EGP'
                currency_out = amount_in.value[-3:]                            
                amount_out = locale.atof(amount_in.value[0:-4])
            elif isinstance(amount_in.value, (int, float)):
                currency_out = amount_in.number_format[-4:-1] # number_format == '0.00 [$USD]'
                amount_out = amount_in.value
            else:
                raise RuntimeError('Wrong type for amount value "{0}" in sheet {1}, row {2}'.format(type(amount_in.value), sheet_in, line_in))
        elif isinstance(amount_in, ReadOnlyCell):
            if isinstance(amount_in.value, str):
                # '235,01'
                amount_out = locale.atof(amount_in.value)
            elif isinstance(amount_in.value, (int, float)):
                amount_out = amount_in.value
            else:
                raise RuntimeError('Wrong type for amount value "{0}" in sheet {1}, row {2}'.format(type(amount_in.value), sheet_in, line_in))
        else:
            raise RuntimeError('Wrong type for amount "{0}" in sheet {1}, row {2}'.format(type(amount_in), sheet_in, line_in))
        amount_out = D(str(amount_out)) # convert to str to keep just the last two decimals

        return sign_out * amount_out, currency_out

    new_page_names = [
        'Datum',
        'ICS-klantnummer',
        'Volgnummer',
        'Bladnummer'
    ]
    page_date = None
    balance_names = [
        "Vorig openstaand saldo",
        "Totaal ontvangen betalingen",
        "Totaal nieuwe uitgaven",
        "Nieuw openstaand saldo"
    ]
    # A line (transaction)
    transaction_names = [
        "Datum transactie",
        "Datum boeking",
        "Omschrijving",
        "Plaats", # Optional together with the next
        "Land", # Optional together with the previous
        "Bedrag in vreemde valuta", # Optional
        "Bedrag in euro's",
        "Bij/Af"
    ]

    current_locale = locale.setlocale(category=locale.LC_ALL) # Save locale
    
    locale.setlocale(category=locale.LC_ALL, locale="Dutch") # Need to parse "05 mei" i.e. "05 may"

    try:
        entries = []
        balances = []
        account = None
        new_page = False
        opening_balance = None
        closing_balance = None
        base_amount_total = 0
        filename = os.path.abspath(filename)
        wb = load_workbook(filename, read_only=True)
        for sheet_i, sheet_name in enumerate(wb.sheetnames, start=1):
            sheet = wb[sheet_name]
            for line_i, row in enumerate(sheet.rows, start=1):
                # Only keep the non null columns
                row = [col for col in row if col.value != None]
                # Handle the two new page rows
                if (sheet_i == 1 and line_i == 2):
                    assert convert_str_to_list(row[0].value, 4) == new_page_names
                    new_page = True
                elif (sheet_i > 1 and line_i == 1):
                    assert [col.value for col in row[0:4]] == new_page_names
                    new_page = True
                elif new_page:
                    new_page = False
                    if page_date is None:
                        page_date, account = convert_str_to_list(row[0].value, 2)
                        page_date = datetime.datetime.strptime(page_date, '%d %B %Y').date()
                # Handle the balance row
                elif (sheet_i == 1 and line_i == 4) or (sheet_i > 1 and line_i == 3):
                    balance_names_actual = None
                    if sheet_i == 1:
                        cols = convert_str_to_list(row[0].value, 12)
                        balance_names_actual = cols[0:4]

                        # opening balance
                        opening_balance, currency = get_amount(sheet_i, line_i, cols[4], cols[5])

                        # closing balance
                        closing_balance, currency = get_amount(sheet_i, line_i, cols[-2], cols[-1])
                        balances.append(
                            RawBalance(
                                account=account,
                                date=page_date,
                                amount=Amount(number=closing_balance, currency=currency),
                                filename=filename,
                                line=((sheet_i-1)*100)+line_i))
                    else:
                        cols0 = convert_str_to_list(row[0].value, 3)
                        cols1 = convert_str_to_list(row[1].value, 1)
                        cols2 = convert_str_to_list(row[2].value, 1)
                        cols3 = convert_str_to_list(row[3].value, 1)
                        # sheet 'Table 3' contains "Nieuw openstaand saldo Af" in fourth column
                        balance_names_actual = [cols0[0], cols1[0], cols2[0], cols3[0].replace(' Af', '')]                                        
                    assert balance_names_actual == balance_names, print("Actual: {0}; Expected: {1}".format(balance_names_actual, balance_names))
                elif (sheet_i == 1 and line_i == 5) or (sheet_i > 1 and line_i == 4):
                    # line with transaction_names
                    continue
                else:
                    # Handle the rest but only for the interesting lines (5, 7 or 8 non-empty columns)
                    #
                    # One strange exception in icscards_balance_wrong.xlsx:
                    #
                    # """
                    # 31 jan          31 jan           IDEAL BETALING, DANK U                                                                                                            861,86         Bij
                    # Uw Card met als laatste vier cijfers 0467
                    # G.J.L.M. PAULISSEN
                    # """
                    #
                    # Another strange exception in icscards_missing.xlsx, Table 2 (page 2):
                    # """
                    # 12 sep         13 sep          SNCF                                               MONTIGNY LE B           FR                                                         6,15          Af
                    # """
                    #
                    if len(row) == 1:
                        maxitems = None
                        line = row[0].value
                        m = re.search('^(.+)\nUw Card met als laatste vier cijfers', line)
                        if m:
                            maxitems = 5
                        else:
                            line = line.strip()
                            m = re.search(' (Bij|Af)$', line)
                            maxitems = 8

                        if m:
                            row = []
                            for column_i, column_value in enumerate(convert_str_to_list(line, maxitems), start=1):
                                row.append(ReadOnlyCell(sheet=sheet, row=line_i, column=column_i, value=column_value))

                    elif len(row) == 6:
                        #
                        # In icscards_missing.xlsx there is something like:
                        #
                        # 20 aug | 21 aug | NETFLIX.COM                                866-579-7172 | NL |  7.99 | Af
                        #
                        # probably due to a later line like:
                        #
                        # 04 sep | 05 sep | NEWREST WAGONS LITS FRANCPARIS | FR | 10.96 | Af
                        #
                        # Solution: create two cells with the first 25 characters and everything thereafter and strip
                        #
                        if DEBUG:
                            breakpoint()
                        new_row = []
                        for column_i, cell in enumerate(row, start=1):
                            if column_i in [1, 2]:
                                new_row.append(cell)
                            elif column_i in [4, 5, 6]:
                                cell.column = cell.column + 1
                                new_row.append(cell)
                            else:
                                assert column_i == 3
                                new_row.append(ReadOnlyCell(sheet=sheet, row=line_i, column=3, value=cell.value[0:25].strip()))
                                new_row.append(ReadOnlyCell(sheet=sheet, row=line_i, column=4, value=cell.value[25:].strip()))
                        row = new_row
                        
                    if not(len(row) == 5 or len(row) == 7 or len(row) == 8):
                        continue
                    
                    # GJP 2020-03-01
                    # Skip transaction date (index 0) since it gives a wrong balance.
                    # Use booking date (index 1) in order to get a correct balance.
                    
                    date = row[1].value
                    try:
                        date = get_date(date)
                    except Exception as e:
                        raise RuntimeError('Invalid date: {0}'.format(date)) from e
                        
                    payee = None
                    narration = None
                    source_desc = None

                    # Add place and country
                    if len(row) >= 7:
                        # payee (2)
                        payee = row[2].value
                        # place (3) and country (4)
                        narration = "{0} ({1})".format(row[3].value, row[4].value)
                        source_desc = "{0}, {1}".format(payee, narration)
                    else:
                        # description (2)
                        narration = row[2].value
                        source_desc = narration

                    # Is there a foreign currency or not (column -3 for 8 columns)?
                    foreign_amount = None
                    foreign_currency = None
                    if len(row) == 8:
                        # foreign_amount always positive as there is no sign info
                        foreign_amount, foreign_currency = get_amount(sheet_i, line_i, row[-3], row[-1], True)

                    # Skip amount in foreign currency
                    base_amount, base_currency = get_amount(sheet_i, line_i, row[-2], row[-1])
                    
                    if base_amount == ZERO:
                        # Skip zero-dollar transactions.
                        # Some banks produce these, e.g. for an annual fee that is waived.
                        continue
                        
                    # In beancount we make a transaction in the foreign currency and have a cost in EUR.
                    # Cost is always positive.
                    amount = None
                    cost = None
                    if False: # forget about cost right now
                        amount=Amount(number=copysign(1, base_amount) * foreign_amount, currency=foreign_currency)
                        cost=Amount(number=abs(base_amount), currency=base_currency)
                    else:
                        amount=Amount(number=base_amount, currency=base_currency)
                        cost=None

                    base_amount_total += base_amount
                        
                    entry = ICScardsEntry(account=account,
                                          date=date,
                                          payee=payee,
                                          narration=narration,
                                          source_desc=source_desc,
                                          amount=amount,
                                          cost=cost,
                                          filename=filename,
                                          line=((sheet_i-1)*100)+line_i)
                    if DEBUG:
                        print(entry)
                    entries.append(entry)

        # No need to sort balances: there is just one
        assert len(balances) == 1, print("Only the balance for the first sheet should be there")
        assert opening_balance + base_amount_total == closing_balance, \
            print("Opening balance ({}) plus the base amount total ({}) should be equal to the closing balance ({})".\
                  format(opening_balance, base_amount_total, closing_balance))
        # No need to sort entries: already sorted by sheet + line
        
        # entries.reverse()
        # entries.sort(key=lambda x: x.line)  # sort by date first, line next
        if DEBUG:
            print(entries)
            print(balances)

    except Exception as e:
        backtrace = "".join(traceback.TracebackException.from_exception(e).format())
        if DEBUG:
            breakpoint()
        raise RuntimeError(backtrace + "\nXLSX file has incorrect format", filename) from e
    finally:
        locale.setlocale(category=locale.LC_ALL, locale=current_locale)

    return entries, balances


def _get_key_from_posting(entry: Transaction, posting: Posting,
                          source_postings: List[Posting], source_desc: str,
                          posting_date: datetime.date):
    del entry
    del source_postings
    return (posting.account, posting_date, posting.units, source_desc)


def _get_key_from_entry(x: ICScardsEntry):
    return (x.account, x.date, x.amount, x.source_desc)


def _make_import_result(icscards_entry: ICScardsEntry) -> ImportResult:
    transaction = Transaction(
        meta=None,
        date=icscards_entry.date,
        flag=FLAG_OKAY,
        payee=icscards_entry.payee,
        narration=icscards_entry.narration,
        tags=EMPTY_SET,
        links=EMPTY_SET,
        postings=[
            Posting(
                account=icscards_entry.account,
                units=icscards_entry.amount,
                cost=icscards_entry.cost,
                price=None,
                flag=None,
                meta=collections.OrderedDict(
                    source_desc=icscards_entry.source_desc,
                    date=icscards_entry.date,
                )),
            Posting(
                account=FIXME_ACCOUNT,
                units=-icscards_entry.amount,
                cost=None,
                price=None,
                flag=None,
                meta=None,
            ),
        ])
    return ImportResult(
        date=icscards_entry.date, info=get_info(icscards_entry), entries=[transaction])

class ICScardsSource(description_based_source.DescriptionBasedSource):
    def __init__(self,
                 filenames: List[str],
                 **kwargs) -> None:
        super().__init__(**kwargs)
        self.filenames = filenames

        # In these entries, account refers to the icscards_id, not the journal account.
        # Balances are in the same file
        self.icscards_entries = []
        self.balances = [] 
        for filename in filenames:
            self.log_status('icscards: loading %s' % filename)
            icscards_entries, balances = load_transactions(filename)
            self.icscards_entries.extend(icscards_entries)
            self.balances.extend(balances)

    def prepare(self, journal: JournalEditor, results: SourceResults) -> None:
        account_to_icscards_id, icscards_id_to_account = description_based_source.get_account_mapping(
            journal.accounts, 'icscards_id')
        missing_accounts = set()  # type: Set[str]

        def get_converted_icscards_entries(entries):
            for raw_icscards_entry in entries:
                account = icscards_id_to_account.get(raw_icscards_entry.account)
                if not account:
                    missing_accounts.add(raw_icscards_entry.account)
                    continue
                match_entry = raw_icscards_entry._replace(account=account)
                yield match_entry

        description_based_source.get_pending_and_invalid_entries(
            raw_entries=get_converted_icscards_entries(self.icscards_entries),
            journal_entries=journal.all_entries,
            account_set=account_to_icscards_id.keys(),
            get_key_from_posting=_get_key_from_posting,
            get_key_from_raw_entry=_get_key_from_entry,
            make_import_result=_make_import_result,
            results=results)

        for icscards_account in missing_accounts:
            results.add_warning(
                'No Beancount account associated with ICScards account %r.' %
                (icscards_account, ))

        for raw_balance in get_converted_icscards_entries(self.balances):
            # GJP 2020-03-01
            # The balance is correct till (excluding) this ICScards date,
            # so it is the beancount definition (at 00:00:00)
            
            # date = raw_balance.date + datetime.timedelta(days=1)
            date = raw_balance.date
            results.add_pending_entry(
                ImportResult(
                    date=date,
                    info=get_info(raw_balance),
                    entries=[
                        Balance(
                            account=raw_balance.account,
                            date=date,
                            meta=None,
                            amount=raw_balance.amount,
                            tolerance=None,
                            diff_amount=None)
                    ]))

    @property
    def name(self):
        return 'icscards'

def load(spec, log_status):
    return ICScardsSource(log_status=log_status, **spec)
